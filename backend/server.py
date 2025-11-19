from fastapi import FastAPI, APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict
import uuid
from datetime import datetime, timezone, date
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# ============= MODELS =============

class Product(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    sell_price: float
    production_cost: float
    category: str  # 'bakso', 'minuman'
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ProductCreate(BaseModel):
    name: str
    sell_price: float
    production_cost: float
    category: str

class Settings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = "settings"
    sewa_harian: float = 150000
    gaji_karyawan_harian: float = 60000
    gaji_owner_harian: float = 50000
    bonus_percentage: float = 0.05
    bonus_max: float = 10000
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SettingsUpdate(BaseModel):
    sewa_harian: Optional[float] = None
    gaji_karyawan_harian: Optional[float] = None
    gaji_owner_harian: Optional[float] = None
    bonus_percentage: Optional[float] = None
    bonus_max: Optional[float] = None

class TransactionItem(BaseModel):
    product_name: str
    quantity: int
    price: float
    production_cost: float

class ExtraItem(BaseModel):
    item_name: str
    quantity: int
    price: float

class PackageTransaction(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str
    package_price: float
    quantity: int = 1
    items: List[TransactionItem]
    extra_items: List[ExtraItem] = []
    extra_items_revenue: float = 0
    total_production_cost: float
    revenue: float
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PackageTransactionCreate(BaseModel):
    date: str
    package_price: float
    quantity: int = 1
    extra_items: List[ExtraItem] = []

class BeverageTransaction(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str
    product_name: str
    quantity: int
    total_price: float
    total_production_cost: float
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BeverageTransactionCreate(BaseModel):
    date: str
    product_name: str
    quantity: int

class DailySummary(BaseModel):
    date: str
    package_revenue: float
    beverage_revenue: float
    total_revenue: float
    package_production_cost: float
    beverage_production_cost: float
    total_production_cost: float
    fixed_costs: float
    net_profit: float
    employee_bonus: float
    package_count: int
    beverage_count: int

class MonthlySummary(BaseModel):
    year: int
    month: int
    total_revenue: float
    total_production_cost: float
    total_fixed_costs: float
    total_net_profit: float
    total_employee_bonus: float
    days_count: int
    daily_summaries: List[DailySummary]

# Define which items carry over to next day
CARRY_OVER_ITEMS = ['bakso_urat', 'bakso_kecil', 'tahu', 'somay']
NON_CARRY_OVER_ITEMS = ['pangsit_malang', 'soun']

class StockItem(BaseModel):
    bakso_urat: int = 0
    bakso_kecil: int = 0
    tahu: int = 0
    somay: int = 0
    pangsit_malang: int = 0
    soun: int = 0

class DailyStock(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str
    stock_brought: StockItem
    stock_remaining: Optional[StockItem] = None
    stock_sold: Optional[StockItem] = None
    revenue_from_stock: Optional[float] = None
    production_cost_from_stock: Optional[float] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DailyStockCreate(BaseModel):
    date: str
    bakso_urat: int
    bakso_kecil: int
    tahu: int
    somay: int
    pangsit_malang: int
    soun: int

class DailyStockRemainingUpdate(BaseModel):
    bakso_urat: int
    bakso_kecil: int
    tahu: int
    somay: int
    pangsit_malang: int
    soun: int

class UnexpectedExpense(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str
    description: str
    amount: float
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UnexpectedExpenseCreate(BaseModel):
    date: str
    description: str
    amount: float

# ============= HELPER FUNCTIONS =============

def calculate_package_composition(package_price: float) -> Dict:
    """
    Hitung komposisi paket berdasarkan harga
    Paket 10.000 = 9 pcs (1 bakso urat, 2 bakso kecil, 1 somay, 1 tahu, 2 pangsit, 1 soun)
    +Rp 1000 = +1 pcs
    Bakso urat = 2 pcs equivalent
    """
    base_price = 10000
    base_pcs = 9
    
    # Hitung total pcs
    price_diff = package_price - base_price
    additional_pcs = price_diff / 1000
    total_pcs = base_pcs + additional_pcs
    
    # Base composition untuk 10.000
    composition = {
        'Bakso urat': 1,
        'Bakso kecil': 2,
        'Somay': 1,
        'Tahu': 1,
        'Pangsit malang': 2,
        'Soun': 1
    }
    
    # Tambah bakso kecil untuk harga lebih tinggi
    if additional_pcs > 0:
        composition['Bakso kecil'] += int(additional_pcs)
    elif additional_pcs < 0:
        # Untuk harga lebih rendah, kurangi dari bakso kecil
        composition['Bakso kecil'] += int(additional_pcs)
        if composition['Bakso kecil'] < 0:
            composition['Bakso kecil'] = 0
    
    return composition

def calculate_production_cost(composition: Dict) -> float:
    """
    Hitung biaya produksi berdasarkan komposisi
    """
    costs = {
        'Bakso urat': 1300,
        'Bakso kecil': 650,
        'Somay': 650,
        'Tahu': 650,
        'Pangsit malang': 650,
        'Soun': 650
    }
    
    total_cost = 0
    for item, quantity in composition.items():
        total_cost += costs.get(item, 0) * quantity
    
    return total_cost

async def get_or_create_settings() -> Settings:
    """Get settings or create default"""
    settings_doc = await db.settings.find_one({"id": "settings"}, {"_id": 0})
    if settings_doc:
        if isinstance(settings_doc.get('updated_at'), str):
            settings_doc['updated_at'] = datetime.fromisoformat(settings_doc['updated_at'])
        return Settings(**settings_doc)
    
    # Create default
    default_settings = Settings()
    doc = default_settings.model_dump()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.settings.insert_one(doc)
    return default_settings

# ============= ROUTES =============

@api_router.get("/")
async def root():
    return {"message": "Bakso Business System API"}

# Products
@api_router.get("/products", response_model=List[Product])
async def get_products():
    products = await db.products.find({}, {"_id": 0}).to_list(1000)
    for p in products:
        if isinstance(p.get('created_at'), str):
            p['created_at'] = datetime.fromisoformat(p['created_at'])
    return products

@api_router.post("/products", response_model=Product)
async def create_product(product: ProductCreate):
    product_obj = Product(**product.model_dump())
    doc = product_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.products.insert_one(doc)
    return product_obj

# Settings
@api_router.get("/settings", response_model=Settings)
async def get_settings():
    return await get_or_create_settings()

@api_router.put("/settings", response_model=Settings)
async def update_settings(updates: SettingsUpdate):
    current = await get_or_create_settings()
    update_data = updates.model_dump(exclude_unset=True)
    
    for key, value in update_data.items():
        setattr(current, key, value)
    
    current.updated_at = datetime.now(timezone.utc)
    doc = current.model_dump()
    doc['updated_at'] = doc['updated_at'].isoformat()
    
    await db.settings.update_one(
        {"id": "settings"},
        {"$set": doc},
        upsert=True
    )
    return current

# Package Transactions
@api_router.post("/transactions/package", response_model=PackageTransaction)
async def create_package_transaction(transaction: PackageTransactionCreate):
    # Calculate composition and cost per package
    composition = calculate_package_composition(transaction.package_price)
    production_cost_per_package = calculate_production_cost(composition)
    
    # Multiply by quantity
    total_production_cost = production_cost_per_package * transaction.quantity
    total_revenue = transaction.package_price * transaction.quantity
    
    # Create transaction items (already multiplied by quantity)
    items = []
    for product_name, qty_per_package in composition.items():
        if qty_per_package > 0:
            item_cost = 1300 if product_name == 'Bakso urat' else 650
            total_qty = qty_per_package * transaction.quantity
            items.append(TransactionItem(
                product_name=product_name,
                quantity=total_qty,
                price=0,  # Harga paket total
                production_cost=item_cost * total_qty
            ))
    
    # Calculate extra items revenue and production cost
    extra_items_revenue = 0
    extra_items_production_cost = 0
    
    for extra in transaction.extra_items:
        # Price is already total for the quantity, don't multiply again
        extra_items_revenue += extra.price
        # Production cost for extra items (quantity is number of pieces)
        if extra.item_name == 'Bakso urat':
            extra_items_production_cost += 1300 * extra.quantity
        else:
            extra_items_production_cost += 650 * extra.quantity
    
    # Add extra items revenue to total
    total_revenue += extra_items_revenue
    total_production_cost += extra_items_production_cost
    
    transaction_obj = PackageTransaction(
        date=transaction.date,
        package_price=transaction.package_price,
        quantity=transaction.quantity,
        items=items,
        extra_items=transaction.extra_items,
        extra_items_revenue=extra_items_revenue,
        total_production_cost=total_production_cost,
        revenue=total_revenue
    )
    
    doc = transaction_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.package_transactions.insert_one(doc)
    
    return transaction_obj

# Beverage Transactions
@api_router.post("/transactions/beverage", response_model=BeverageTransaction)
async def create_beverage_transaction(transaction: BeverageTransactionCreate):
    # Get product info
    beverages = {
        'Teh rosela': {'price': 5000, 'cost': 3000},
        'Es teh manis': {'price': 3000, 'cost': 2000}
    }
    
    beverage_info = beverages.get(transaction.product_name)
    if not beverage_info:
        raise HTTPException(status_code=400, detail="Produk minuman tidak ditemukan")
    
    total_price = beverage_info['price'] * transaction.quantity
    total_cost = beverage_info['cost'] * transaction.quantity
    
    transaction_obj = BeverageTransaction(
        date=transaction.date,
        product_name=transaction.product_name,
        quantity=transaction.quantity,
        total_price=total_price,
        total_production_cost=total_cost
    )
    
    doc = transaction_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.beverage_transactions.insert_one(doc)
    
    return transaction_obj

# Daily Summary
@api_router.get("/daily-summary/{date}", response_model=DailySummary)
async def get_daily_summary(date: str):
    # Get stock data for the date (untuk paket bakso)
    stock_data = await db.daily_stocks.find_one({"date": date}, {"_id": 0})
    
    package_revenue = 0
    package_cost = 0
    package_count = 0
    
    if stock_data and stock_data.get('stock_remaining'):
        # Revenue dari stok management (paket bakso)
        package_revenue = stock_data.get('revenue_from_stock', 0)
        package_cost = stock_data.get('production_cost_from_stock', 0)
        package_count = 1  # Indicate stock is managed
    
    # Get all beverage transactions for date
    beverage_txns = await db.beverage_transactions.find({"date": date}, {"_id": 0}).to_list(1000)
    
    beverage_revenue = sum(t['total_price'] for t in beverage_txns)
    beverage_cost = sum(t['total_production_cost'] for t in beverage_txns)
    beverage_count = len(beverage_txns)
    
    # Get unexpected expenses for the date
    unexpected_expenses = await db.unexpected_expenses.find({"date": date}, {"_id": 0}).to_list(1000)
    total_unexpected_expenses = sum(e['amount'] for e in unexpected_expenses)
    
    # Get settings for fixed costs
    settings = await get_or_create_settings()
    fixed_costs = settings.sewa_harian + settings.gaji_karyawan_harian + settings.gaji_owner_harian
    
    # Calculate totals (include unexpected expenses)
    total_revenue = package_revenue + beverage_revenue
    total_production_cost = package_cost + beverage_cost
    net_profit = total_revenue - total_production_cost - fixed_costs - total_unexpected_expenses
    
    # Calculate employee bonus
    employee_bonus = 0
    if net_profit > 0:
        bonus_calculated = net_profit * settings.bonus_percentage
        employee_bonus = min(bonus_calculated, settings.bonus_max)
    
    return DailySummary(
        date=date,
        package_revenue=package_revenue,
        beverage_revenue=beverage_revenue,
        total_revenue=total_revenue,
        package_production_cost=package_cost,
        beverage_production_cost=beverage_cost,
        total_production_cost=total_production_cost,
        fixed_costs=fixed_costs,
        net_profit=net_profit,
        employee_bonus=employee_bonus,
        package_count=package_count,
        beverage_count=beverage_count
    )

# Monthly Summary
@api_router.get("/monthly-summary/{year}/{month}", response_model=MonthlySummary)
async def get_monthly_summary(year: int, month: int):
    # Get all dates in month that have transactions
    package_txns = await db.package_transactions.find({}, {"_id": 0, "date": 1}).to_list(10000)
    beverage_txns = await db.beverage_transactions.find({}, {"_id": 0, "date": 1}).to_list(10000)
    
    # Extract unique dates for the month
    dates = set()
    for t in package_txns + beverage_txns:
        txn_date = t['date']
        if txn_date.startswith(f"{year}-{month:02d}"):
            dates.add(txn_date)
    
    # Get daily summaries for all dates
    daily_summaries = []
    total_revenue = 0
    total_production_cost = 0
    total_fixed_costs = 0
    total_net_profit = 0
    total_employee_bonus = 0
    
    for date_str in sorted(dates):
        summary = await get_daily_summary(date_str)
        daily_summaries.append(summary)
        total_revenue += summary.total_revenue
        total_production_cost += summary.total_production_cost
        total_fixed_costs += summary.fixed_costs
        total_net_profit += summary.net_profit
        total_employee_bonus += summary.employee_bonus
    
    return MonthlySummary(
        year=year,
        month=month,
        total_revenue=total_revenue,
        total_production_cost=total_production_cost,
        total_fixed_costs=total_fixed_costs,
        total_net_profit=total_net_profit,
        total_employee_bonus=total_employee_bonus,
        days_count=len(dates),
        daily_summaries=daily_summaries
    )

# Get package transactions for a date
@api_router.get("/transactions/package/{date}", response_model=List[PackageTransaction])
async def get_package_transactions(date: str):
    transactions = await db.package_transactions.find({"date": date}, {"_id": 0}).to_list(1000)
    for t in transactions:
        if isinstance(t.get('created_at'), str):
            t['created_at'] = datetime.fromisoformat(t['created_at'])
    return transactions

# Get beverage transactions for a date
@api_router.get("/transactions/beverage/{date}", response_model=List[BeverageTransaction])
async def get_beverage_transactions(date: str):
    transactions = await db.beverage_transactions.find({"date": date}, {"_id": 0}).to_list(1000)
    for t in transactions:
        if isinstance(t.get('created_at'), str):
            t['created_at'] = datetime.fromisoformat(t['created_at'])
    return transactions

# Delete transaction
@api_router.delete("/transactions/package/{transaction_id}")
async def delete_package_transaction(transaction_id: str):
    result = await db.package_transactions.delete_one({"id": transaction_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")
    return {"message": "Transaksi berhasil dihapus"}

@api_router.delete("/transactions/beverage/{transaction_id}")
async def delete_beverage_transaction(transaction_id: str):
    result = await db.beverage_transactions.delete_one({"id": transaction_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Transaksi tidak ditemukan")
    return {"message": "Transaksi berhasil dihapus"}

# Stock Management
@api_router.post("/stock/initial", response_model=DailyStock)
async def create_initial_stock(stock: DailyStockCreate):
    # Check if stock already exists for this date
    existing = await db.daily_stocks.find_one({"date": stock.date}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Stok untuk tanggal ini sudah ada")
    
    stock_brought = StockItem(
        bakso_urat=stock.bakso_urat,
        bakso_kecil=stock.bakso_kecil,
        tahu=stock.tahu,
        somay=stock.somay,
        pangsit_malang=stock.pangsit_malang,
        soun=stock.soun
    )
    
    stock_obj = DailyStock(
        date=stock.date,
        stock_brought=stock_brought
    )
    
    doc = stock_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    await db.daily_stocks.insert_one(doc)
    
    return stock_obj

@api_router.put("/stock/remaining/{date}", response_model=DailyStock)
async def update_remaining_stock(date: str, remaining: DailyStockRemainingUpdate):
    # Get existing stock
    existing = await db.daily_stocks.find_one({"date": date}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Stok awal belum diinput untuk tanggal ini")
    
    if isinstance(existing.get('created_at'), str):
        existing['created_at'] = datetime.fromisoformat(existing['created_at'])
    if isinstance(existing.get('updated_at'), str):
        existing['updated_at'] = datetime.fromisoformat(existing['updated_at'])
    
    stock_obj = DailyStock(**existing)
    
    # Set remaining stock
    stock_obj.stock_remaining = StockItem(**remaining.model_dump())
    
    # Calculate sold = brought - remaining
    # Pangsit dan soun tidak dibawa pulang, jadi sisa harus 0
    sold = {
        'bakso_urat': stock_obj.stock_brought.bakso_urat - remaining.bakso_urat,
        'bakso_kecil': stock_obj.stock_brought.bakso_kecil - remaining.bakso_kecil,
        'tahu': stock_obj.stock_brought.tahu - remaining.tahu,
        'somay': stock_obj.stock_brought.somay - remaining.somay,
        'pangsit_malang': stock_obj.stock_brought.pangsit_malang - remaining.pangsit_malang,
        'soun': stock_obj.stock_brought.soun - remaining.soun
    }
    
    stock_obj.stock_sold = StockItem(**sold)
    
    # Calculate revenue from sold stock
    # Harga jual: Bakso urat Rp 2000, lainnya Rp 1000
    revenue = (
        sold['bakso_urat'] * 2000 +
        sold['bakso_kecil'] * 1000 +
        sold['tahu'] * 1000 +
        sold['somay'] * 1000 +
        sold['pangsit_malang'] * 1000 +
        sold['soun'] * 1000
    )
    stock_obj.revenue_from_stock = revenue
    
    # Calculate production cost from sold stock
    # Biaya produksi: Bakso urat Rp 1300, lainnya Rp 650
    production_cost = (
        sold['bakso_urat'] * 1300 +
        sold['bakso_kecil'] * 650 +
        sold['tahu'] * 650 +
        sold['somay'] * 650 +
        sold['pangsit_malang'] * 650 +
        sold['soun'] * 650
    )
    stock_obj.production_cost_from_stock = production_cost
    
    stock_obj.updated_at = datetime.now(timezone.utc)
    
    # Update in database
    doc = stock_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    
    await db.daily_stocks.update_one(
        {"date": date},
        {"$set": doc}
    )
    
    return stock_obj

@api_router.get("/stock/{date}", response_model=DailyStock)
async def get_daily_stock(date: str):
    stock = await db.daily_stocks.find_one({"date": date}, {"_id": 0})
    if not stock:
        raise HTTPException(status_code=404, detail="Stok tidak ditemukan untuk tanggal ini")
    
    if isinstance(stock.get('created_at'), str):
        stock['created_at'] = datetime.fromisoformat(stock['created_at'])
    if isinstance(stock.get('updated_at'), str):
        stock['updated_at'] = datetime.fromisoformat(stock['updated_at'])
    
    return DailyStock(**stock)

@api_router.put("/stock/initial/{date}", response_model=DailyStock)
async def update_initial_stock(date: str, stock: DailyStockCreate):
    # Get existing stock
    existing = await db.daily_stocks.find_one({"date": date}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Stok tidak ditemukan")
    
    if isinstance(existing.get('created_at'), str):
        existing['created_at'] = datetime.fromisoformat(existing['created_at'])
    if isinstance(existing.get('updated_at'), str):
        existing['updated_at'] = datetime.fromisoformat(existing['updated_at'])
    
    stock_obj = DailyStock(**existing)
    
    # Update stock brought
    stock_obj.stock_brought = StockItem(
        bakso_urat=stock.bakso_urat,
        bakso_kecil=stock.bakso_kecil,
        tahu=stock.tahu,
        somay=stock.somay,
        pangsit_malang=stock.pangsit_malang,
        soun=stock.soun
    )
    
    # If remaining exists, recalculate sold and revenue
    if stock_obj.stock_remaining:
        remaining = stock_obj.stock_remaining
        sold = {
            'bakso_urat': stock.bakso_urat - remaining.bakso_urat,
            'bakso_kecil': stock.bakso_kecil - remaining.bakso_kecil,
            'tahu': stock.tahu - remaining.tahu,
            'somay': stock.somay - remaining.somay,
            'pangsit_malang': stock.pangsit_malang - remaining.pangsit_malang,
            'soun': stock.soun - remaining.soun
        }
        
        stock_obj.stock_sold = StockItem(**sold)
        
        # Recalculate revenue
        revenue = (
            sold['bakso_urat'] * 2000 +
            sold['bakso_kecil'] * 1000 +
            sold['tahu'] * 1000 +
            sold['somay'] * 1000 +
            sold['pangsit_malang'] * 1000 +
            sold['soun'] * 1000
        )
        stock_obj.revenue_from_stock = revenue
        
        # Recalculate production cost
        production_cost = (
            sold['bakso_urat'] * 1300 +
            sold['bakso_kecil'] * 650 +
            sold['tahu'] * 650 +
            sold['somay'] * 650 +
            sold['pangsit_malang'] * 650 +
            sold['soun'] * 650
        )
        stock_obj.production_cost_from_stock = production_cost
    
    stock_obj.updated_at = datetime.now(timezone.utc)
    
    # Update in database
    doc = stock_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    
    await db.daily_stocks.update_one(
        {"date": date},
        {"$set": doc}
    )
    
    return stock_obj

@api_router.delete("/stock/{date}")
async def delete_daily_stock(date: str):
    result = await db.daily_stocks.delete_one({"date": date})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Stok tidak ditemukan")
    return {"message": "Stok berhasil dihapus"}

# Export Endpoints
@api_router.get("/export/daily/{date}")
async def export_daily_report(date: str):
    """Export laporan harian ke Excel"""
    # Get daily summary
    summary = await get_daily_summary(date)
    
    # Get stock data
    stock_data = await db.daily_stocks.find_one({"date": date}, {"_id": 0})
    
    # Get beverage transactions
    beverage_txns = await db.beverage_transactions.find({"date": date}, {"_id": 0}).to_list(1000)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Laporan {date}"
    
    # Styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = f"LAPORAN HARIAN BAKSO BUSINESS"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f"Tanggal: {date}"
    ws['A2'].font = Font(bold=True)
    ws.merge_cells('A2:D2')
    
    # Summary Section
    row = 4
    ws[f'A{row}'] = "RINGKASAN KEUANGAN"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = Font(color="FFFFFF", bold=True)
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    ws[f'A{row}'] = "Total Pendapatan"
    ws[f'B{row}'] = f"Rp {summary.total_revenue:,.0f}"
    row += 1
    ws[f'A{row}'] = "Biaya Produksi"
    ws[f'B{row}'] = f"Rp {summary.total_production_cost:,.0f}"
    row += 1
    ws[f'A{row}'] = "Biaya Tetap"
    ws[f'B{row}'] = f"Rp {summary.fixed_costs:,.0f}"
    row += 1
    ws[f'A{row}'] = "Laba Bersih"
    ws[f'B{row}'] = f"Rp {summary.net_profit:,.0f}"
    ws[f'B{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = "Bonus Karyawan"
    ws[f'B{row}'] = f"Rp {summary.employee_bonus:,.0f}"
    ws[f'B{row}'].font = Font(bold=True, color="00B050")
    
    # Stock Section
    if stock_data:
        row += 3
        ws[f'A{row}'] = "DATA STOK BAKSO"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].font = Font(color="FFFFFF", bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        headers = ['Item', 'Bawaan', 'Sisa', 'Terjual']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        if stock_data.get('stock_remaining'):
            stock_items = [
                ('Bakso Urat', 'bakso_urat'),
                ('Bakso Kecil', 'bakso_kecil'),
                ('Tahu', 'tahu'),
                ('Somay', 'somay'),
                ('Pangsit Malang', 'pangsit_malang'),
                ('Soun', 'soun')
            ]
            
            for label, key in stock_items:
                row += 1
                ws[f'A{row}'] = label
                ws[f'B{row}'] = stock_data['stock_brought'][key]
                ws[f'C{row}'] = stock_data['stock_remaining'][key]
                ws[f'D{row}'] = stock_data['stock_sold'][key]
            
            row += 1
            ws[f'A{row}'] = "Pendapatan dari Stok"
            ws[f'B{row}'] = f"Rp {stock_data.get('revenue_from_stock', 0):,.0f}"
            ws[f'B{row}'].font = Font(bold=True)
            ws.merge_cells(f'B{row}:D{row}')
    
    # Beverage Section
    if beverage_txns:
        row += 3
        ws[f'A{row}'] = "TRANSAKSI MINUMAN"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].font = Font(color="FFFFFF", bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        headers = ['Produk', 'Jumlah', 'Total Harga', 'Biaya Produksi']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for txn in beverage_txns:
            row += 1
            ws[f'A{row}'] = txn['product_name']
            ws[f'B{row}'] = txn['quantity']
            ws[f'C{row}'] = f"Rp {txn['total_price']:,.0f}"
            ws[f'D{row}'] = f"Rp {txn['total_production_cost']:,.0f}"
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Laporan_Harian_{date}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/export/monthly/{year}/{month}")
async def export_monthly_report(year: int, month: int):
    """Export laporan bulanan ke Excel"""
    # Get monthly summary
    summary = await get_monthly_summary(year, month)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Laporan {year}-{month:02d}"
    
    # Styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = f"LAPORAN BULANAN BAKSO BUSINESS"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    ws['A2'] = f"Periode: {year}-{month:02d}"
    ws['A2'].font = Font(bold=True)
    ws.merge_cells('A2:E2')
    
    # Summary Section
    row = 4
    ws[f'A{row}'] = "RINGKASAN BULANAN"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = Font(color="FFFFFF", bold=True)
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    ws[f'A{row}'] = "Total Pendapatan"
    ws[f'B{row}'] = f"Rp {summary.total_revenue:,.0f}"
    row += 1
    ws[f'A{row}'] = "Total Biaya Produksi"
    ws[f'B{row}'] = f"Rp {summary.total_production_cost:,.0f}"
    row += 1
    ws[f'A{row}'] = "Total Biaya Tetap"
    ws[f'B{row}'] = f"Rp {summary.total_fixed_costs:,.0f}"
    row += 1
    ws[f'A{row}'] = "Total Laba Bersih"
    ws[f'B{row}'] = f"Rp {summary.total_net_profit:,.0f}"
    ws[f'B{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = "Total Bonus Karyawan"
    ws[f'B{row}'] = f"Rp {summary.total_employee_bonus:,.0f}"
    ws[f'B{row}'].font = Font(bold=True, color="00B050")
    row += 1
    ws[f'A{row}'] = "Jumlah Hari Operasional"
    ws[f'B{row}'] = summary.days_count
    
    # Daily Details
    row += 3
    ws[f'A{row}'] = "DETAIL HARIAN"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = header_fill
    ws[f'A{row}'].font = Font(color="FFFFFF", bold=True)
    ws.merge_cells(f'A{row}:E{row}')
    
    row += 1
    headers = ['Tanggal', 'Pendapatan', 'Biaya Produksi', 'Laba Bersih', 'Bonus']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    for daily in summary.daily_summaries:
        row += 1
        ws[f'A{row}'] = daily.date
        ws[f'B{row}'] = f"Rp {daily.total_revenue:,.0f}"
        ws[f'C{row}'] = f"Rp {daily.total_production_cost:,.0f}"
        ws[f'D{row}'] = f"Rp {daily.net_profit:,.0f}"
        ws[f'E{row}'] = f"Rp {daily.employee_bonus:,.0f}"
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Laporan_Bulanan_{year}_{month:02d}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Unexpected Expenses
@api_router.post("/unexpected-expenses", response_model=UnexpectedExpense)
async def create_unexpected_expense(expense: UnexpectedExpenseCreate):
    expense_obj = UnexpectedExpense(**expense.model_dump())
    doc = expense_obj.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.unexpected_expenses.insert_one(doc)
    return expense_obj

@api_router.get("/unexpected-expenses/{date}", response_model=List[UnexpectedExpense])
async def get_unexpected_expenses(date: str):
    expenses = await db.unexpected_expenses.find({"date": date}, {"_id": 0}).to_list(1000)
    for e in expenses:
        if isinstance(e.get('created_at'), str):
            e['created_at'] = datetime.fromisoformat(e['created_at'])
    return expenses

@api_router.delete("/unexpected-expenses/{expense_id}")
async def delete_unexpected_expense(expense_id: str):
    result = await db.unexpected_expenses.delete_one({"id": expense_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Pengeluaran tidak ditemukan")
    return {"message": "Pengeluaran berhasil dihapus"}

@api_router.get("/export/stock/{date}")
async def export_stock_data(date: str):
    """Export data stok ke Excel"""
    stock_data = await db.daily_stocks.find_one({"date": date}, {"_id": 0})
    
    if not stock_data:
        raise HTTPException(status_code=404, detail="Data stok tidak ditemukan")
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Stok {date}"
    
    # Styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Title
    ws['A1'] = f"DATA STOK BAKSO - {date}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    # Headers
    row = 3
    headers = ['Item', 'Bawaan', 'Sisa', 'Terjual']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    stock_items = [
        ('Bakso Urat', 'bakso_urat'),
        ('Bakso Kecil', 'bakso_kecil'),
        ('Tahu', 'tahu'),
        ('Somay', 'somay'),
        ('Pangsit Malang', 'pangsit_malang'),
        ('Soun', 'soun')
    ]
    
    for label, key in stock_items:
        row += 1
        ws[f'A{row}'] = label
        ws[f'B{row}'] = stock_data['stock_brought'][key]
        ws[f'C{row}'] = stock_data.get('stock_remaining', {}).get(key, '-')
        ws[f'D{row}'] = stock_data.get('stock_sold', {}).get(key, '-')
    
    # Summary
    if stock_data.get('revenue_from_stock'):
        row += 2
        ws[f'A{row}'] = "Pendapatan dari Stok:"
        ws[f'B{row}'] = f"Rp {stock_data['revenue_from_stock']:,.0f}"
        ws[f'B{row}'].font = Font(bold=True, color="00B050")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Data_Stok_{date}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()